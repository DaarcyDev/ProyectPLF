from tkinter import StringVar, ttk
from tkinter.font import Font
import customtkinter
from PIL import Image, ImageTk
from openpyxl import load_workbook
import random

# Cargar el archivo Excel
workbookMaterias = load_workbook(filename='Boris.xlsx')
sheetMaterias = workbookMaterias.active

workbookMaestros = load_workbook(filename='Mestros.xlsx')
sheetMaestros = workbookMaestros.active

# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("System")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("dark-blue")

window = customtkinter.CTk()

# window.geometry("400x780")
window.title("Proyecto Tkinter")
# window.resizable(0,0)
# window.grid_columnconfigure(1, weight=1)
# window.grid_rowconfigure(0, weight=0)


class Actions:
    btn_font = customtkinter.CTkFont(family="Roboto Cn", size=25)

    indexFramePic = customtkinter.CTkFrame(window)

    indexFrame = customtkinter.CTkFrame(window)
    indexFrame2 = customtkinter.CTkFrame(window)

    aboutFrame = customtkinter.CTkFrame(window)
    aboutFrame2 = customtkinter.CTkFrame(window)

    aboutFrame3 = customtkinter.CTkFrame(window)

    menuFrame = customtkinter.CTkFrame(window)

    # productsBox = customtkinter.Treeview(window,height=19, columns=("#0","#1"), selectmode="extended")
    productsBox = ttk.Treeview(window, height=35, columns=(
        "#0", "#1"), selectmode="extended")

    def destroy():
        for widget in Actions.aboutFrame.winfo_children():
            widget.destroy()
        for widget in Actions.aboutFrame2.winfo_children():
            widget.destroy()
        for widget in Actions.aboutFrame3.winfo_children():
            widget.destroy()
        for widget in Actions.indexFrame.winfo_children():
            widget.destroy()
        for widget in Actions.indexFrame2.winfo_children():
            widget.destroy()
        for widget in Actions.indexFramePic.winfo_children():
            widget.destroy()
        for widget in Actions.productsBox.winfo_children():
            widget.destroy()
        Actions.productsBox.delete(*Actions.productsBox.get_children())

    def Generar():
        Actions.destroy()

        Actions.productsBox.pack(pady=40, padx=60, anchor="center")

        search_var = StringVar()

        Actions.aboutFrame3.pack(
            pady=20, padx=60, expand=False, anchor="center")

        aboutLabel1 = customtkinter.CTkLabel(
            master=Actions.aboutFrame3, justify=customtkinter.LEFT, text="Buscador:", font=Actions.btn_font)
        aboutLabel1.pack(pady=10, padx=10)

        generarInput = customtkinter.CTkEntry(
            master=Actions.aboutFrame3, justify=customtkinter.LEFT, font=Actions.btn_font, textvariable=search_var)
        generarInput.pack(pady=10, padx=10)

        Actions.productsBox['columns'] = (
            "Maestros", 'Materias', 'Creditos', 'Grupos', 'Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', "Check")

        Actions.productsBox.column("#0", width=0)
        Actions.productsBox.column("Maestros", width=350)
        Actions.productsBox.column("Materias", width=450)
        Actions.productsBox.column("Creditos", width=100)
        Actions.productsBox.column("Grupos", width=100)
        Actions.productsBox.column("Lunes", width=100)
        Actions.productsBox.column("Martes", width=100)
        Actions.productsBox.column("Miercoles", width=100)
        Actions.productsBox.column("Jueves", width=100)
        Actions.productsBox.column("Viernes", width=100)
        Actions.productsBox.column("Check", width=100)

        Actions.productsBox.heading("#0", text="")
        Actions.productsBox.heading("Maestros", text="Maestros")
        Actions.productsBox.heading("Materias", text="Materias")
        Actions.productsBox.heading("Creditos", text="Creditos")
        Actions.productsBox.heading("Grupos", text="Grupos")
        Actions.productsBox.heading("Lunes", text="Lunes")
        Actions.productsBox.heading("Martes", text="Martes")
        Actions.productsBox.heading("Miercoles", text="Miercoes")
        Actions.productsBox.heading("Jueves", text="Jueves")
        Actions.productsBox.heading("Viernes", text="Viernes")
        Actions.productsBox.heading("Check", text="Check")

        maestros_list = list(sheetMaestros.iter_rows(
            min_row=2, values_only=True))
        maestros = random.sample(maestros_list, 89)

        materias_list = list(sheetMaterias.iter_rows(
            min_row=2, values_only=True))

        for maestro in maestros:

            test1 = ""
            test2 = ""
            test3 = ""
            test4 = ""
            test5 = ""

            materias = random.sample(materias_list, 5)

            for materia in materias:

                if test1 == "":
                    test1 = materia
                    if (test1[3] == None):
                        test1 = list(test1)
                        test1[3] = "-"
                        test1 = tuple(test1)
                elif test2 == "":
                    test2 = materia
                    if (test2[3] == None):
                        test2 = list(test2)
                        test2[3] = "-"
                        test2 = tuple(test2)
                    if test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test1[4], test2[4], test1[5],
                              test2[5], test1[6], test2[6], test1[7], test2[7])

                        while test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test2 = materia
                            if test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test2 = materia
                                if (test2[3] == None):
                                    test2 = list(test2)
                                    test2[3] = "-"
                                    test2 = tuple(test2)
                                if (test2[4] == None):
                                    test2 = list(test2)
                                    test2[4] = "-"
                                    test2 = tuple(test2)
                                if (test2[5] == None):
                                    test2 = list(test2)
                                    test2[5] = "-"
                                    test2 = tuple(test2)
                                if (test2[6] == None):
                                    test2 = list(test2)
                                    test2[6] = "-"
                                    test2 = tuple(test2)
                                if (test2[7] == None):
                                    test2 = list(test2)
                                    test2[7] = "-"
                                    test2 = tuple(test2)
                                print("arreglando test2")

                            else:
                                print("arreglado test2")
                                print(test1[3], test2[3], test1[4], test2[4], test1[5],
                                      test2[5], test1[6], test2[6], test1[7], test2[7])
                                break

                    else:
                        print("son diferentes N1")
                elif test3 == "":
                    test3 = materia
                    if (test3[3] == None):
                        test3 = list(test3)
                        test3[3] = "-"
                        test3 = tuple(test3)
                    if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3])
                        while test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test3 = materia
                            if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test3 = materia
                                if (test3[3] == None):
                                    test3 = list(test3)
                                    test3[3] = "--"
                                    test3 = tuple(test3)
                                if (test3[4] == None):
                                    test3 = list(test3)
                                    test3[4] = "--"
                                    test3 = tuple(test3)
                                if (test3[5] == None):
                                    test3 = list(test3)
                                    test3[5] = "--"
                                    test3 = tuple(test3)
                                if (test3[6] == None):
                                    test3 = list(test3)
                                    test3[6] = "--"
                                    test3 = tuple(test3)
                                if (test3[7] == None):
                                    test3 = list(test3)
                                    test3[7] = "--"
                                    test3 = tuple(test3)
                                print("arreglando test3")

                            else:
                                print("arreglado test3")
                                print(test1[3], test2[3], test3[3])
                                break

                    else:
                        print("son diferentes N2")
                elif test4 == "":
                    test4 = materia
                    if (test4[3] == None):
                        test4 = list(test4)
                        test4[3] = "-"
                        test4 = tuple(test4)
                    if test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3])
                        while test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test4 = materia
                            if test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test4 = materia
                                if (test4[3] == None):
                                    test4 = list(test4)
                                    test4[3] = "---"
                                    test4 = tuple(test4)
                                if (test4[4] == None):
                                    test4 = list(test4)
                                    test4[4] = "---"
                                    test4 = tuple(test4)
                                if (test4[5] == None):
                                    test4 = list(test4)
                                    test4[5] = "---"
                                    test4 = tuple(test4)
                                if (test4[6] == None):
                                    test4 = list(test4)
                                    test4[6] = "---"
                                    test4 = tuple(test4)
                                if (test4[7] == None):
                                    test4 = list(test4)
                                    test4[7] = "---"
                                    test4 = tuple(test4)
                                print("arreglando test4")

                            else:

                                print("arreglado test4")
                                print(test1[3], test2[3], test3[3], test4[3])
                                break

                    else:
                        print("son diferentes N3")
                elif test5 == "":
                    test5 = materia
                    if (test5[3] == None):
                        test5 = list(test5)
                        test5[3] = "-"
                        test5 = tuple(test5)
                    if test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3], test5[3])
                        while test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7] or str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                            materia = random.sample(materias_list, k=1)[0]
                            test5 = materia
                            if test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                                if (test5[3] == None):
                                    test5 = list(test5)
                                    test5[3] = "----"
                                    test5 = tuple(test5)
                                if (test5[4] == None):
                                    test5 = list(test5)
                                    test5[4] = "----"
                                    test5 = tuple(test5)
                                if (test5[5] == None):
                                    test5 = list(test5)
                                    test5[5] = "----"
                                    test5 = tuple(test5)
                                if (test5[6] == None):
                                    test5 = list(test5)
                                    test5[6] = "----"
                                    test5 = tuple(test5)
                                if (test5[7] == None):
                                    test5 = list(test5)
                                    test5[7] = "----"
                                    test5 = tuple(test5)
                                print("arreglando test5")

                            else:

                                print("arreglado test5")
                                print(test1[3], test2[3],
                                      test3[3], test4[3], test5[3])
                                break
                            if str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                                print("10 horas por dia")
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                            else:
                                print("menos de 10 horas por dia")

                    else:
                        print("son diferentes N4")


            for val in materias:
                # print("----------\n"+str(test1[3])+"--"+str(test1[4])+"--"+str(test1[5])+"--"+str(test1[6])+"--"+str(test1[7])+"\n"+str(test2[3])+"--"+str(test2[4])+"--"+str(test2[5])+"--"+str(test2[6])+"--"+str(test2[7])+"\n"+str(test3[3])+"--"+str(test3[4])+"--"+str(test3[5])+"--"+str(test3[6])+"--"+str(test3[7])+"\n"+str(test4[3])+"--"+str(test4[4])+"--"+str(test4[5])+"--"+str(test4[6])+"--"+str(test4[7])+"\n"+str(test5[3])+"--"+str(test5[4])+"--"+str(test5[5])+"--"+str(test5[6])+"--"+str(test5[7])+"\n")
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=("", "", "", "", "", "", "", "", "", ""))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test1[0]), str(test1[1]), str(test1[2]), str(test1[3]), str(test1[4]), str(test1[5]), str(test1[6]), str(test1[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test2[0]), str(test2[1]), str(test2[2]), str(test2[3]), str(test2[4]), str(test2[5]), str(test2[6]), str(test2[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test3[0]), str(test3[1]), str(test3[2]), str(test3[3]), str(test3[4]), str(test3[5]), str(test3[6]), str(test3[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test4[0]), str(test4[1]), str(test4[2]), str(test4[3]), str(test4[4]), str(test4[5]), str(test4[6]), str(test4[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test5[0]), str(test5[1]), str(test5[2]), str(test5[3]), str(test5[4]), str(test5[5]), str(test5[6]), str(test5[7]), "1"))
                break

        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.Eliminar(), text="eliminar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="right")
        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.Crear(), text="crear", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="right")
        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.menu(), text="regresar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")
        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.CrearUno(), text="Crear Uno", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")
        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.Buscador(search_var.get()), text="Buscar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")

        Actions.indexFrame.pack_forget()
        Actions.indexFrame2.pack_forget()
        Actions.aboutFrame.pack_forget()
        Actions.aboutFrame2.pack_forget()
        Actions.indexFramePic.pack_forget()

    def Eliminar():
        for record in Actions.productsBox.get_children():
            Actions.productsBox.delete(record)

    def Crear():
        maestros_list = list(sheetMaestros.iter_rows(
            min_row=2, values_only=True))
        maestros = random.sample(maestros_list, 89)

        materias_list = list(sheetMaterias.iter_rows(
            min_row=2, values_only=True))

        for maestro in maestros:
            
            test1 = ""
            test2 = ""
            test3 = ""
            test4 = ""
            test5 = ""
            
            materias = random.sample(materias_list,5)

            for materia in materias:

                if test1 == "":
                    test1 = materia
                    if(test1[3]==None):
                        test1 = list(test1)
                        test1[3] = "-"
                        test1 = tuple(test1)
                elif test2 == "":
                    test2 = materia
                    if(test2[3]==None):
                        test2 = list(test2)
                        test2[3] = "-"
                        test2 = tuple(test2)
                    if  test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                        
                        while test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test2 = materia
                            if test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test2 = materia
                                if(test2[3]==None):
                                    test2 = list(test2)
                                    test2[3] = "-"
                                    test2 = tuple(test2)
                                if(test2[4]==None):
                                    test2 = list(test2)
                                    test2[4] = "-"
                                    test2 = tuple(test2)
                                if(test2[5]==None):
                                    test2 = list(test2)
                                    test2[5] = "-"
                                    test2 = tuple(test2)
                                if(test2[6]==None):
                                    test2 = list(test2)
                                    test2[6] = "-"
                                    test2 = tuple(test2)
                                if(test2[7]==None):
                                    test2 = list(test2)
                                    test2[7] = "-"
                                    test2 = tuple(test2)
                                print("arreglando test2")
                                
                                
                            else:
                                print("arreglado test2")
                                print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                                break
                                
                            
                        
                    else:
                        print("son diferentes N1")
                elif test3 == "":
                    test3 = materia
                    if(test3[3]==None):
                        test3 = list(test3)
                        test3[3] = "-"
                        test3 = tuple(test3)
                    if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3])
                        while test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test3 = materia
                            if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test3 = materia
                                if(test3[3]==None):
                                    test3 = list(test3)
                                    test3[3] = "--"
                                    test3 = tuple(test3)
                                if(test3[4]==None):
                                    test3 = list(test3)
                                    test3[4] = "--"
                                    test3 = tuple(test3)
                                if(test3[5]==None):
                                    test3 = list(test3)
                                    test3[5] = "--"
                                    test3 = tuple(test3)
                                if(test3[6]==None):
                                    test3 = list(test3)
                                    test3[6] = "--"
                                    test3 = tuple(test3)
                                if(test3[7]==None):
                                    test3 = list(test3)
                                    test3[7] = "--"
                                    test3 = tuple(test3)
                                print("arreglando test3")
                                
                                
                            else:
                                print("arreglado test3")
                                print(test1[3], test2[3], test3[3])
                                break
                                
                    else:
                        print("son diferentes N2")
                elif test4 == "":
                    test4 = materia
                    if(test4[3]==None):
                        test4 = list(test4)
                        test4[3] = "-"
                        test4 = tuple(test4)
                    if test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3])
                        while test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test4 = materia
                            if  test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test4 = materia
                                if(test4[3]==None):
                                    test4 = list(test4)
                                    test4[3] = "---"
                                    test4 = tuple(test4)
                                if(test4[4]==None):
                                    test4 = list(test4)
                                    test4[4] = "---"
                                    test4 = tuple(test4)
                                if(test4[5]==None):
                                    test4 = list(test4)
                                    test4[5] = "---"
                                    test4 = tuple(test4)
                                if(test4[6]==None):
                                    test4 = list(test4)
                                    test4[6] = "---"
                                    test4 = tuple(test4)
                                if(test4[7]==None):
                                    test4 = list(test4)
                                    test4[7] = "---"
                                    test4 = tuple(test4)
                                print("arreglando test4")
                                
                            else:
                                
                                print("arreglado test4")
                                print(test1[3], test2[3], test3[3], test4[3])
                                break
                                
                    else:
                        print("son diferentes N3")
                elif test5 == "":
                    test5 = materia
                    if(test5[3]==None):
                        test5 = list(test5)
                        test5[3] = "-"
                        test5 = tuple(test5)
                    if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3], test5[3])
                        while test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7] or str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                            materia = random.sample(materias_list, k=1)[0]
                            test5 = materia
                            if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                                if(test5[3]==None):
                                    test5 = list(test5)
                                    test5[3] = "----"
                                    test5 = tuple(test5)
                                if(test5[4]==None):
                                    test5 = list(test5)
                                    test5[4] = "----"
                                    test5 = tuple(test5)
                                if(test5[5]==None):
                                    test5 = list(test5)
                                    test5[5] = "----"
                                    test5 = tuple(test5)
                                if(test5[6]==None):
                                    test5 = list(test5)
                                    test5[6] = "----"
                                    test5 = tuple(test5)
                                if(test5[7]==None):
                                    test5 = list(test5)
                                    test5[7] = "----"
                                    test5 = tuple(test5)
                                print("arreglando test5")
                                
                            else:
                                
                                print("arreglado test5")
                                print(test1[3], test2[3], test3[3], test4[3], test5[3])
                                break
                            if str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                                print("10 horas por dia")
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                            else:
                                print("menos de 10 horas por dia")

                                
                    else:
                        print("son diferentes N4")

            for val in materias:
                # print("----------\n"+str(test1[3])+"--"+str(test1[4])+"--"+str(test1[5])+"--"+str(test1[6])+"--"+str(test1[7])+"\n"+str(test2[3])+"--"+str(test2[4])+"--"+str(test2[5])+"--"+str(test2[6])+"--"+str(test2[7])+"\n"+str(test3[3])+"--"+str(test3[4])+"--"+str(test3[5])+"--"+str(test3[6])+"--"+str(test3[7])+"\n"+str(test4[3])+"--"+str(test4[4])+"--"+str(test4[5])+"--"+str(test4[6])+"--"+str(test4[7])+"\n"+str(test5[3])+"--"+str(test5[4])+"--"+str(test5[5])+"--"+str(test5[6])+"--"+str(test5[7])+"\n")
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=("", "", "", "", "", "", "", "", "", ""))

                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test1[0]), str(test1[1]), str(test1[2]), str(test1[3]), str(test1[4]), str(test1[5]), str(test1[6]), str(test1[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test2[0]), str(test2[1]), str(test2[2]), str(test2[3]), str(test2[4]), str(test2[5]), str(test2[6]), str(test2[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test3[0]), str(test3[1]), str(test3[2]), str(test3[3]), str(test3[4]), str(test3[5]), str(test3[6]), str(test3[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test4[0]), str(test4[1]), str(test4[2]), str(test4[3]), str(test4[4]), str(test4[5]), str(test4[6]), str(test4[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test5[0]), str(test5[1]), str(test5[2]), str(test5[3]), str(test5[4]), str(test5[5]), str(test5[6]), str(test5[7]), "1"))
                break

    def CrearUno():
        maestros_list = list(sheetMaestros.iter_rows(
            min_row=2, values_only=True))
        maestros = random.sample(maestros_list, 1)

        materias_list = list(sheetMaterias.iter_rows(
            min_row=2, values_only=True))

        for maestro in maestros:
            
            test1 = ""
            test2 = ""
            test3 = ""
            test4 = ""
            test5 = ""
            
            materias = random.sample(materias_list,5)

            for materia in materias:

                if test1 == "":
                    test1 = materia
                    if(test1[3]==None):
                        test1 = list(test1)
                        test1[3] = "-"
                        test1 = tuple(test1)
                elif test2 == "":
                    test2 = materia
                    if(test2[3]==None):
                        test2 = list(test2)
                        test2[3] = "-"
                        test2 = tuple(test2)
                    if  test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                        
                        while test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test2 = materia
                            if test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test2 = materia
                                if(test2[3]==None):
                                    test2 = list(test2)
                                    test2[3] = "-"
                                    test2 = tuple(test2)
                                if(test2[4]==None):
                                    test2 = list(test2)
                                    test2[4] = "-"
                                    test2 = tuple(test2)
                                if(test2[5]==None):
                                    test2 = list(test2)
                                    test2[5] = "-"
                                    test2 = tuple(test2)
                                if(test2[6]==None):
                                    test2 = list(test2)
                                    test2[6] = "-"
                                    test2 = tuple(test2)
                                if(test2[7]==None):
                                    test2 = list(test2)
                                    test2[7] = "-"
                                    test2 = tuple(test2)
                                print("arreglando test2")
                                
                                
                            else:
                                print("arreglado test2")
                                print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                                break
                                
                            
                        
                    else:
                        print("son diferentes N1")
                elif test3 == "":
                    test3 = materia
                    if(test3[3]==None):
                        test3 = list(test3)
                        test3[3] = "-"
                        test3 = tuple(test3)
                    if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3])
                        while test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test3 = materia
                            if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test3 = materia
                                if(test3[3]==None):
                                    test3 = list(test3)
                                    test3[3] = "--"
                                    test3 = tuple(test3)
                                if(test3[4]==None):
                                    test3 = list(test3)
                                    test3[4] = "--"
                                    test3 = tuple(test3)
                                if(test3[5]==None):
                                    test3 = list(test3)
                                    test3[5] = "--"
                                    test3 = tuple(test3)
                                if(test3[6]==None):
                                    test3 = list(test3)
                                    test3[6] = "--"
                                    test3 = tuple(test3)
                                if(test3[7]==None):
                                    test3 = list(test3)
                                    test3[7] = "--"
                                    test3 = tuple(test3)
                                print("arreglando test3")
                                
                                
                            else:
                                print("arreglado test3")
                                print(test1[3], test2[3], test3[3])
                                break
                                
                    else:
                        print("son diferentes N2")
                elif test4 == "":
                    test4 = materia
                    if(test4[3]==None):
                        test4 = list(test4)
                        test4[3] = "-"
                        test4 = tuple(test4)
                    if test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3])
                        while test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test4 = materia
                            if  test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test4 = materia
                                if(test4[3]==None):
                                    test4 = list(test4)
                                    test4[3] = "---"
                                    test4 = tuple(test4)
                                if(test4[4]==None):
                                    test4 = list(test4)
                                    test4[4] = "---"
                                    test4 = tuple(test4)
                                if(test4[5]==None):
                                    test4 = list(test4)
                                    test4[5] = "---"
                                    test4 = tuple(test4)
                                if(test4[6]==None):
                                    test4 = list(test4)
                                    test4[6] = "---"
                                    test4 = tuple(test4)
                                if(test4[7]==None):
                                    test4 = list(test4)
                                    test4[7] = "---"
                                    test4 = tuple(test4)
                                print("arreglando test4")
                                
                            else:
                                
                                print("arreglado test4")
                                print(test1[3], test2[3], test3[3], test4[3])
                                break
                                
                    else:
                        print("son diferentes N3")
                elif test5 == "":
                    test5 = materia
                    if(test5[3]==None):
                        test5 = list(test5)
                        test5[3] = "-"
                        test5 = tuple(test5)
                    if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3], test5[3])
                        while test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7] or str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                            materia = random.sample(materias_list, k=1)[0]
                            test5 = materia
                            if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                                if(test5[3]==None):
                                    test5 = list(test5)
                                    test5[3] = "----"
                                    test5 = tuple(test5)
                                if(test5[4]==None):
                                    test5 = list(test5)
                                    test5[4] = "----"
                                    test5 = tuple(test5)
                                if(test5[5]==None):
                                    test5 = list(test5)
                                    test5[5] = "----"
                                    test5 = tuple(test5)
                                if(test5[6]==None):
                                    test5 = list(test5)
                                    test5[6] = "----"
                                    test5 = tuple(test5)
                                if(test5[7]==None):
                                    test5 = list(test5)
                                    test5[7] = "----"
                                    test5 = tuple(test5)
                                print("arreglando test5")
                                
                            else:
                                
                                print("arreglado test5")
                                print(test1[3], test2[3], test3[3], test4[3], test5[3])
                                break
                            if str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                                print("10 horas por dia")
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                            else:
                                print("menos de 10 horas por dia")

                                
                    else:
                        print("son diferentes N4")
            for val in materias:

                # print("----------\n"+str(test1[3])+"--"+str(test1[4])+"--"+str(test1[5])+"--"+str(test1[6])+"--"+str(test1[7])+"\n"+str(test2[3])+"--"+str(test2[4])+"--"+str(test2[5])+"--"+str(test2[6])+"--"+str(test2[7])+"\n"+str(test3[3])+"--"+str(test3[4])+"--"+str(test3[5])+"--"+str(test3[6])+"--"+str(test3[7])+"\n"+str(test4[3])+"--"+str(test4[4])+"--"+str(test4[5])+"--"+str(test4[6])+"--"+str(test4[7])+"\n"+str(test5[3])+"--"+str(test5[4])+"--"+str(test5[5])+"--"+str(test5[6])+"--"+str(test5[7])+"\n")
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=("", "", "", "", "", "", "", "", "", ""))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test1[0]), str(test1[1]), str(test1[2]), str(test1[3]), str(test1[4]), str(test1[5]), str(test1[6]), str(test1[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test2[0]), str(test2[1]), str(test2[2]), str(test2[3]), str(test2[4]), str(test2[5]), str(test2[6]), str(test2[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test3[0]), str(test3[1]), str(test3[2]), str(test3[3]), str(test3[4]), str(test3[5]), str(test3[6]), str(test3[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test4[0]), str(test4[1]), str(test4[2]), str(test4[3]), str(test4[4]), str(test4[5]), str(test4[6]), str(test4[7]), "1"))
                Actions.productsBox.insert(
                    "", customtkinter.END, text="", values=(str(maestro[0]), str(test5[0]), str(test5[1]), str(test5[2]), str(test5[3]), str(test5[4]), str(test5[5]), str(test5[6]), str(test5[7]), "1"))
                break

    def Buscador(search_term):
        items = Actions.productsBox.get_children()
        for item in items:
            values = Actions.productsBox.item(item, "values")
            if search_term.lower() in " ".join(map(str, values)).lower():
                Actions.productsBox.selection_add(item)
            else:
                Actions.productsBox.selection_remove(item)

    def Horarios():
        Actions.destroy()

        Actions.indexFrame2.pack(
            pady=20, padx=60, expand=True, anchor="center")

        Actions.productsBox.pack(pady=40, padx=60, anchor="center")

        Actions.aboutFrame3.pack(
            pady=20, padx=60, expand=False, anchor="center")

        aboutLabel1 = customtkinter.CTkLabel(
            master=Actions.indexFrame2, justify=customtkinter.LEFT, text="Horarios", font=Actions.btn_font)
        aboutLabel1.pack(pady=10, padx=10)

        Actions.productsBox['columns'] = (
            'Materias', 'Creditos', 'Grupos', 'Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes')

        Actions.productsBox.column("#0", width=0)
        Actions.productsBox.column("Materias", width=250)
        Actions.productsBox.column("Creditos", width=150)
        Actions.productsBox.column("Grupos", width=150)
        Actions.productsBox.column("Lunes", width=150)
        Actions.productsBox.column("Martes", width=150)
        Actions.productsBox.column("Miercoles", width=150)
        Actions.productsBox.column("Jueves", width=150)
        Actions.productsBox.column("Viernes", width=150)
        # Actions.productsBox.column("player_states",width=80)
        # Actions.productsBox.column("player_city",width=80)

        Actions.productsBox.heading("#0", text="")
        Actions.productsBox.heading("Materias", text="Materias")
        Actions.productsBox.heading("Creditos", text="Creditos")
        Actions.productsBox.heading("Grupos", text="Grupos")
        Actions.productsBox.heading("Lunes", text="Lunes")
        Actions.productsBox.heading("Martes", text="Martes")
        Actions.productsBox.heading("Miercoles", text="Miercoes")
        Actions.productsBox.heading("Jueves", text="Jueves")
        Actions.productsBox.heading("Viernes", text="Viernes")

        for row in sheetMaterias.iter_rows(min_row=2, values_only=True):
            Actions.productsBox.insert(
                "", customtkinter.END, text="", values=tuple(row))

        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.menu(), text="regresar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")

        Actions.indexFrame.pack_forget()
        Actions.indexFrame2.pack_forget()
        Actions.aboutFrame.pack_forget()
        Actions.aboutFrame2.pack_forget()
        Actions.indexFramePic.pack_forget()

    def Maestros():
        Actions.destroy()

        Actions.productsBox.pack(pady=40, padx=60)

        Actions.aboutFrame3.pack(
            pady=20, padx=60, expand=False, anchor="center")

        Actions.productsBox['columns'] = ('maestros', 'HPlazas', 'HAsignadas')

        Actions.productsBox.column("#0", width=0)
        Actions.productsBox.column("maestros", width=350)
        Actions.productsBox.column("HPlazas", width=150)
        Actions.productsBox.column("HAsignadas", width=150)

        Actions.productsBox.heading("#0", text="")
        Actions.productsBox.heading("maestros", text="Maestros")
        Actions.productsBox.heading("HPlazas", text="Horas de plaza")
        Actions.productsBox.heading("HAsignadas", text="Horas Asignadas")

        for row in sheetMaestros.iter_rows(min_row=2, values_only=True):
            Actions.productsBox.insert(
                "", customtkinter.END, text="", values=tuple(row))

        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame3, width=300, height=50, command=lambda: Actions.menu(), text="regresar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")

        Actions.indexFrame.pack_forget()
        Actions.indexFrame2.pack_forget()
        Actions.aboutFrame.pack_forget()
        Actions.aboutFrame2.pack_forget()
        Actions.indexFramePic.pack_forget()
        Actions.menuFrame.pack_forget()
        # Actions.productsBox.pack_forget()

    def menu():
        Actions.destroy()

        Actions.aboutFrame.pack(pady=20, padx=60, fill="both", expand=True)
        Actions.aboutFrame2.pack(
            pady=20, padx=60, expand=True, anchor="center")

        aboutButton1 = customtkinter.CTkButton(
            Actions.aboutFrame, width=300, height=50, command=lambda: Actions.Maestros(), text="Maestros", font=Actions.btn_font)
        aboutButton1.pack(pady=10, padx=10)

        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame, width=300, height=50, command=lambda: Actions.Horarios(), text="Horarios", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10)

        aboutButton3 = customtkinter.CTkButton(
            Actions.aboutFrame, width=300, height=50, command=lambda: Actions.Generar(), text="Generar", font=Actions.btn_font)
        aboutButton3.pack(pady=10, padx=10)

        aboutButton4 = customtkinter.CTkButton(
            Actions.aboutFrame2, width=300, height=50, command=lambda: Actions.index(), text="regresar", font=Actions.btn_font)
        aboutButton4.pack(pady=10, padx=10, side="left")

        Actions.indexFrame.pack_forget()
        Actions.indexFrame2.pack_forget()
        Actions.aboutFrame3.pack_forget()

        Actions.indexFramePic.pack_forget()
        Actions.productsBox.pack_forget()

    def about():
        Actions.destroy()

        Actions.aboutFrame.pack(pady=20, padx=60, fill="both", expand=True)
        Actions.aboutFrame2.pack(
            pady=20, padx=60, expand=True, anchor="center")

        imagen = Image.open("about.jpg")
        nuevo_tamano = (500, 800)
        imagen = imagen.resize(nuevo_tamano)
        imagen_tk = ImageTk.PhotoImage(imagen)

        label = customtkinter.CTkLabel(
            Actions.aboutFrame, image=imagen_tk, text="")
        label.pack(side="left")

        aboutLabel1 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Nombre:", font=Actions.btn_font)
        aboutLabel1.pack(pady=10, padx=10)

        aboutLabel2 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Liam Alvarez Peralta", font=Actions.btn_font)
        aboutLabel2.pack(pady=10, padx=10,)

        aboutLabel3 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Lugar:", font=Actions.btn_font)
        aboutLabel3.pack(pady=10, padx=10,)

        aboutLabel4 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Zacatepec, Morelos", font=Actions.btn_font)
        aboutLabel4.pack(pady=10, padx=10,)

        aboutLabel5 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Fecha:", font=Actions.btn_font)
        aboutLabel5.pack(pady=10, padx=10,)

        aboutLabel6 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="28-Febrero-2023", font=Actions.btn_font)
        aboutLabel6.pack(pady=10, padx=10,)

        aboutLabel7 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Instituto Tecnologico de Zacatepec", font=Actions.btn_font)
        aboutLabel7.pack(pady=10, padx=10,)

        aboutLabel8 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Proyecto:", font=Actions.btn_font)
        aboutLabel8.pack(pady=10, padx=10,)

        aboutLabel9 = customtkinter.CTkLabel(
            master=Actions.aboutFrame, justify=customtkinter.LEFT, text="Asignacion de maestros a grupos usando PLR", font=Actions.btn_font)
        aboutLabel9.pack(pady=10, padx=10,)

        aboutButton2 = customtkinter.CTkButton(
            Actions.aboutFrame2, width=300, height=50, command=lambda: Actions.index(), text="regresar", font=Actions.btn_font)
        aboutButton2.pack(pady=10, padx=10, side="left")

        Actions.indexFrame.pack_forget()
        Actions.indexFrame2.pack_forget()
        Actions.indexFramePic.pack_forget()

    def exit():
        window.destroy()

    def index():

        Actions.indexFramePic.pack(
            pady=20, padx=60, expand=True, fill="both", anchor="n")
        Actions.indexFrame.pack(pady=20, padx=60, fill="both", expand=True)
        Actions.indexFrame2.pack(
            pady=20, padx=60, expand=True, anchor="center")
        Actions.indexFramePic.pack(
            pady=20, padx=60, expand=False, anchor="w")

        imagen = Image.open("itz.png")
        nuevo_tamano = (300, 300)
        imagen = imagen.resize(nuevo_tamano)
        imagen_tk = ImageTk.PhotoImage(imagen)

        label = customtkinter.CTkLabel(
            Actions.indexFramePic, image=imagen_tk, text="")
        label.pack(pady=10, padx=10, side="left")

        imagen = Image.open("tnm.jpg")
        nuevo_tamano = (300, 300)
        imagen = imagen.resize(nuevo_tamano)
        imagen_tk = ImageTk.PhotoImage(imagen)

        label = customtkinter.CTkLabel(
            Actions.indexFramePic, image=imagen_tk, text="")
        label.pack(pady=10, padx=10, side="right")

        imagen = Image.open("isc.jpg")
        nuevo_tamano = (300, 300)
        imagen = imagen.resize(nuevo_tamano)
        imagen_tk = ImageTk.PhotoImage(imagen)

        label = customtkinter.CTkLabel(
            Actions.indexFramePic, image=imagen_tk, text="")
        label.pack(pady=10, padx=10, side="left")

        indexLabel1 = customtkinter.CTkLabel(
            master=Actions.indexFrame, justify=customtkinter.LEFT, text="Tecnologico Nacional de Mexico", font=Actions.btn_font)
        indexLabel1.pack(pady=100, padx=10)

        indexLabel2 = customtkinter.CTkLabel(
            master=Actions.indexFrame, justify=customtkinter.LEFT, text="Instituto Tecnologico de Zacatepec", font=Actions.btn_font)
        indexLabel2.pack(pady=30, padx=10)

        indexLabel3 = customtkinter.CTkLabel(
            master=Actions.indexFrame, justify=customtkinter.LEFT, text="Ingenieria en Sistemas Computacionales", font=Actions.btn_font)
        indexLabel3.pack(pady=30, padx=10)

        indexLabel4 = customtkinter.CTkLabel(
            master=Actions.indexFrame, justify=customtkinter.LEFT, text="Programacion Logica y Funcional", font=Actions.btn_font)
        indexLabel4.pack(pady=30, padx=10)

        indexButton1 = customtkinter.CTkButton(
            Actions.indexFrame2, width=300, height=50, command=lambda: Actions.menu(), text="Menu", font=Actions.btn_font)
        indexButton1.pack(pady=10, padx=10, side="left")

        indexButton2 = customtkinter.CTkButton(
            Actions.indexFrame2, width=300, height=50, command=lambda: Actions.about(), text="Acerca de", font=Actions.btn_font)
        indexButton2.pack(pady=10, padx=10, side="left")

        indexButton3 = customtkinter.CTkButton(
            Actions.indexFrame2, width=300, height=50, command=lambda: Actions.exit(), text="Salir", font=Actions.btn_font)
        indexButton3.pack(pady=10, padx=10, side="left")

        Actions.aboutFrame.pack_forget()
        Actions.aboutFrame2.pack_forget()


Actions.index()
window.mainloop()
