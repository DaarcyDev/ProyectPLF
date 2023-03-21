import random
from openpyxl import load_workbook

# Cargar el archivo Excel
workbookMaterias = load_workbook(filename='Boris.xlsx')
sheetMaterias = workbookMaterias.active

workbookMaestros = load_workbook(filename='Mestros.xlsx')
sheetMaestros = workbookMaestros.active

def Generar():
        maestros_list = list(sheetMaestros.iter_rows(min_row=2, values_only=True))
        maestros = random.sample(maestros_list, 89)



        
        materias_list = list(sheetMaterias.iter_rows(min_row=2, values_only=True))

        for maestro in maestros:
            
            test1 = ""
            test2 = ""
            test3 = ""
            test4 = ""
            test5 = ""
            
            materias = random.sample(materias_list,5)

            for materia in materias:

                if test1 == "":
                    test1 = materia
                    if(test1[3]==None):
                        test1 = list(test1)
                        test1[3] = "-"
                        test1 = tuple(test1)
                elif test2 == "":
                    test2 = materia
                    if(test2[3]==None):
                        test2 = list(test2)
                        test2[3] = "-"
                        test2 = tuple(test2)
                    if  test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                        
                        while test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                            materia = random.sample(materias_list, k=1)[0]
                            test2 = materia
                            if test2[3] == test1[3] or test2[4] == test1[4] or test2[5] == test1[5] or test2[6] == test1[6] or test2[7] == test1[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test2 = materia
                                if(test2[3]==None):
                                    test2 = list(test2)
                                    test2[3] = "-"
                                    test2 = tuple(test2)
                                if(test2[4]==None):
                                    test2 = list(test2)
                                    test2[4] = "-"
                                    test2 = tuple(test2)
                                if(test2[5]==None):
                                    test2 = list(test2)
                                    test2[5] = "-"
                                    test2 = tuple(test2)
                                if(test2[6]==None):
                                    test2 = list(test2)
                                    test2[6] = "-"
                                    test2 = tuple(test2)
                                if(test2[7]==None):
                                    test2 = list(test2)
                                    test2[7] = "-"
                                    test2 = tuple(test2)
                                print("arreglando test2")
                                
                                
                            else:
                                print("arreglado test2")
                                print(test1[3], test2[3], test1[4], test2[4], test1[5], test2[5], test1[6], test2[6], test1[7], test2[7])
                                break
                                
                            
                        
                    else:
                        print("son diferentes N1")
                elif test3 == "":
                    test3 = materia
                    if(test3[3]==None):
                        test3 = list(test3)
                        test3[3] = "-"
                        test3 = tuple(test3)
                    if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3])
                        while test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test3 = materia
                            if test1[3] == test3[3] or test2[3] == test3[3] or test1[4] == test3[4] or test2[4] == test3[4] or test1[5] == test3[5] or test2[5] == test3[5] or test1[6] == test3[6] or test2[6] == test3[6] or test1[7] == test3[7] or test2[7] == test3[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test3 = materia
                                if(test3[3]==None):
                                    test3 = list(test3)
                                    test3[3] = "--"
                                    test3 = tuple(test3)
                                if(test3[4]==None):
                                    test3 = list(test3)
                                    test3[4] = "--"
                                    test3 = tuple(test3)
                                if(test3[5]==None):
                                    test3 = list(test3)
                                    test3[5] = "--"
                                    test3 = tuple(test3)
                                if(test3[6]==None):
                                    test3 = list(test3)
                                    test3[6] = "--"
                                    test3 = tuple(test3)
                                if(test3[7]==None):
                                    test3 = list(test3)
                                    test3[7] = "--"
                                    test3 = tuple(test3)
                                print("arreglando test3")
                                
                                
                            else:
                                print("arreglado test3")
                                print(test1[3], test2[3], test3[3])
                                break
                                
                    else:
                        print("son diferentes N2")
                elif test4 == "":
                    test4 = materia
                    if(test4[3]==None):
                        test4 = list(test4)
                        test4[3] = "-"
                        test4 = tuple(test4)
                    if test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3])
                        while test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                            materia = random.sample(materias_list, k=1)[0]
                            test4 = materia
                            if  test1[3] == test4[3] or test2[3] == test4[3] or test3[3] == test4[3] or test1[4] == test4[4] or test2[4] == test4[4] or test3[4] == test4[4] or test1[5] == test4[5] or test2[5] == test4[5] or test3[5] == test4[5] or test1[6] == test4[6] or test2[6] == test4[6] or test3[6] == test4[6] or test1[7] == test4[7] or test2[7] == test4[7] or test3[7] == test4[7] :
                                materia = random.sample(materias_list, k=1)[0]
                                test4 = materia
                                if(test4[3]==None):
                                    test4 = list(test4)
                                    test4[3] = "---"
                                    test4 = tuple(test4)
                                if(test4[4]==None):
                                    test4 = list(test4)
                                    test4[4] = "---"
                                    test4 = tuple(test4)
                                if(test4[5]==None):
                                    test4 = list(test4)
                                    test4[5] = "---"
                                    test4 = tuple(test4)
                                if(test4[6]==None):
                                    test4 = list(test4)
                                    test4[6] = "---"
                                    test4 = tuple(test4)
                                if(test4[7]==None):
                                    test4 = list(test4)
                                    test4[7] = "---"
                                    test4 = tuple(test4)
                                print("arreglando test4")
                                
                            else:
                                
                                print("arreglado test4")
                                print(test1[3], test2[3], test3[3], test4[3])
                                break
                                
                    else:
                        print("son diferentes N3")
                elif test5 == "":
                    test5 = materia
                    if(test5[3]==None):
                        test5 = list(test5)
                        test5[3] = "-"
                        test5 = tuple(test5)
                    if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                        print("son iguales")
                        print(test1[3], test2[3], test3[3], test4[3], test5[3])
                        while test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7] or str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                            materia = random.sample(materias_list, k=1)[0]
                            test5 = materia
                            if  test1[3] == test5[3] or test2[3] == test5[3] or test3[3] == test5[3] or test4[3] == test5[3] or test1[4] == test5[4] or test2[4] == test5[4] or test3[4] == test5[4] or test4[4] == test5[4] or test1[5] == test5[5] or test2[5] == test5[5] or test3[5] == test5[5] or test4[5] == test5[5] or test1[6] == test5[6] or test2[6] == test5[6] or test3[6] == test5[6] or test4[6] == test5[6] or test1[7] == test5[7] or test2[7] == test5[7] or test3[7] == test5[7] or test4[7] == test5[7]:
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                                if(test5[3]==None):
                                    test5 = list(test5)
                                    test5[3] = "----"
                                    test5 = tuple(test5)
                                if(test5[4]==None):
                                    test5 = list(test5)
                                    test5[4] = "----"
                                    test5 = tuple(test5)
                                if(test5[5]==None):
                                    test5 = list(test5)
                                    test5[5] = "----"
                                    test5 = tuple(test5)
                                if(test5[6]==None):
                                    test5 = list(test5)
                                    test5[6] = "----"
                                    test5 = tuple(test5)
                                if(test5[7]==None):
                                    test5 = list(test5)
                                    test5[7] = "----"
                                    test5 = tuple(test5)
                                print("arreglando test5")
                                
                            else:
                                
                                print("arreglado test5")
                                print(test1[3], test2[3], test3[3], test4[3], test5[3])
                                break
                            if str(test1[3])[0].isdigit() and str(test2[3])[0].isdigit() and str(test3[3])[0].isdigit() and str(test4[3])[0].isdigit() and str(test5[3])[0].isdigit() or str(test1[4])[0].isdigit() and str(test2[4])[0].isdigit() and str(test3[4])[0].isdigit() and str(test4[4])[0].isdigit() and str(test5[4])[0].isdigit() or str(test1[5])[0].isdigit() and str(test2[5])[0].isdigit() and str(test3[5])[0].isdigit() and str(test4[5])[0].isdigit() and str(test5[5])[0].isdigit() or str(test1[6])[0].isdigit() and str(test2[6])[0].isdigit() and str(test3[6])[0].isdigit() and str(test4[6])[0].isdigit() and str(test5[6])[0].isdigit() or str(test1[7])[0].isdigit() and str(test2[7])[0].isdigit() and str(test3[7])[0].isdigit() and str(test4[7])[0].isdigit() and str(test5[7])[0].isdigit():
                                print("10 horas por dia")
                                materia = random.sample(materias_list, k=1)[0]
                                test5 = materia
                            else:
                                print("menos de 10 horas por dia")

                                
                    else:
                        print("son diferentes N4")
                        
            for val in sheetMaterias:
                print("------------------------------")
                print(str(maestro[0]),str(test1[1]),str(test1[2]),str(test1[3]),str(test1[4]),str(test1[5]),str(test1[6]),str(test1[7])+"\n"+
                    str(maestro[0]),str(test2[1]),str(test2[2]),str(test2[3]),str(test2[4]),str(test2[5]),str(test2[6]),str(test2[7])+"\n"+
                    str(maestro[0]),str(test3[1]),str(test3[2]),str(test3[3]),str(test3[4]),str(test3[5]),str(test3[6]),str(test3[7])+"\n"+
                    str(maestro[0]),str(test4[1]),str(test4[2]),str(test4[3]),str(test4[4]),str(test4[5]),str(test4[6]),str(test4[7])+"\n"+
                    str(maestro[0]),str(test5[1]),str(test5[2]),str(test5[3]),str(test5[4]),str(test5[5]),str(test5[6]),str(test5[7])+"\n")
                break
Generar()